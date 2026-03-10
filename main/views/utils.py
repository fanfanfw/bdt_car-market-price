"""
Utility functions and helpers for views
"""
from functools import lru_cache
import random
import re
from datetime import date, datetime, timedelta

from django.http import HttpResponse
from django.utils import timezone
from decouple import config

from ..models import (
    MileageConfiguration, BrandCategory, PriceTier, VerifiedPhone, CalculationLog,
    VehicleConditionCategory
)
from ..api_client import get_price_estimation, get_car_records, get_car_detail, APIError


def _normalize_phone_e164_like(phone: str) -> str:
    """
    Normalize to a simple +<digits> form for consistent matching.
    This is not a full E.164 validator; it only strips non-digits.
    """
    digits = re.sub(r"\D", "", phone or "")
    return f"+{digits}" if digits else ""


@lru_cache(maxsize=1)
def get_otp_bypass_phones() -> set[str]:
    """
    Return a set of normalized phone numbers that bypass OTP.

    Configure via env var OTP_BYPASS_PHONE.
    Example:
      OTP_BYPASS_PHONE=+60123456789,+6281234567890
    """
    raw = config("OTP_BYPASS_PHONE", default="") or ""
    if not raw.strip():
        return set()

    # Split by commas, whitespace, or semicolons.
    parts = [p.strip() for p in re.split(r"[,\s;]+", raw) if p and p.strip()]
    return {p for p in (_normalize_phone_e164_like(x) for x in parts) if p}


def is_otp_bypass_phone(phone: str) -> bool:
    """Check whether a phone number is configured to bypass OTP."""
    normalized = _normalize_phone_e164_like(phone)
    return bool(normalized) and normalized in get_otp_bypass_phones()


def get_mileage_config():
    """Get the mileage configuration"""
    try:
        return MileageConfiguration.objects.first()
    except MileageConfiguration.DoesNotExist:
        # Return default values if no config exists
        return type('obj', (object,), {
            'threshold_percent': 10.0,
            'reduction_percent': 2.0,
            'max_reduction_cap': 15.0,
            'layer2_max_cap': 70.0,
        })


def get_car_statistics(
    brand,
    model,
    variant,
    year,
    user_mileage=None,
    condition_assessments=None,
    selected_condition_details=None,
):
    """Get car statistics including average mileage and price with condition assessments using FastAPI"""
    try:
        # Get mileage configuration
        mileage_config = get_mileage_config()

        # Get car statistics from FastAPI
        try:
            estimation_data = get_price_estimation(
                brand=brand,
                model=model,
                variant=variant,
                year=year,
                mileage=user_mileage
            )

            print(f"DEBUG: FastAPI response for {brand} {model} {variant} {year}: {estimation_data}")

            # Extract statistics from FastAPI response
            stats = estimation_data.get('statistics', {})
            if not stats:
                # If no statistics key, try to extract from the response directly
                if 'estimated_price' in estimation_data:
                    stats = {
                        'average_price': estimation_data.get('price_range', {}).get('avg', estimation_data.get('estimated_price', 0)),
                        'average_mileage': 100000,  # Default value since FastAPI doesn't provide this
                        'data_count': estimation_data.get('sample_size', 1)
                    }
                else:
                    print(f"DEBUG: No valid price data in FastAPI response: {estimation_data}")
                    return None

            avg_mileage = stats.get('average_mileage', 100000)  # Default fallback
            avg_price = stats.get('average_price', 0)
            total_data = stats.get('data_count', estimation_data.get('sample_size', 0))
            raw_price_range = estimation_data.get('price_range', {}) if isinstance(estimation_data, dict) else {}

            if total_data == 0:
                print(f"DEBUG: No data found for {brand} {model} {variant} {year}")
                return None

        except APIError as e:
            print(f"DEBUG: FastAPI error: {e}")
            # Fallback if FastAPI fails - return None to indicate no data
            return None

        result = {
            'brand_norm': brand,
            'model_norm': model,
            'variant_norm': variant,
            'year': year,
            'rata_rata_mileage_bulat': avg_mileage,
            'rata_rata_price_bulat': avg_price,
            'total_data': total_data,
            'sample_size': total_data,
        }

        # Calculate price adjustments with dynamic 2-layer system
        avg_mileage = float(avg_mileage)
        avg_price = float(avg_price)

        # Layer 1: Mileage-based reduction
        layer1_reduction = 0
        mileage_diff_percent = 0

        if user_mileage is not None:
            user_mileage = float(user_mileage)
            if user_mileage > avg_mileage:
                mileage_diff_percent = ((user_mileage - avg_mileage) / avg_mileage) * 100
                threshold = float(mileage_config.threshold_percent)
                reduction_per_threshold = float(mileage_config.reduction_percent)
                layer1_reduction = (mileage_diff_percent / threshold) * reduction_per_threshold
                # Apply cap
                layer1_reduction = min(layer1_reduction, float(mileage_config.max_reduction_cap))
            else:
                mileage_diff_percent = ((user_mileage - avg_mileage) / avg_mileage) * 100

        # Layer 2: Condition assessment reduction + Brand Category Auto-Detection
        layer2_reduction = 0
        condition_breakdown = {}
        brand_category_reduction = 0
        brand_category_info = None

        # Auto-detect brand category reduction
        try:
            brand_category_mapping = BrandCategory.objects.select_related('category').get(brand=brand)
            # Get the reduction percentage directly from the category model
            brand_category_reduction = float(brand_category_mapping.category.reduction_percentage)
            brand_category_info = {
                'brand': brand,
                'category': brand_category_mapping.category.name,
                'reduction': brand_category_reduction
            }
        except BrandCategory.DoesNotExist:
            # Brand not classified - use 0% reduction
            brand_category_reduction = 0
            brand_category_info = {
                'brand': brand,
                'category': 'Unclassified',
                'reduction': 0,
                'warning': 'Brand not classified - admin should classify this brand'
            }

        # Auto-detect price tier reduction
        price_tier_reduction = 0
        price_tier_info = None

        try:
            price_tier = PriceTier.get_tier_for_price(avg_price)
            if price_tier:
                price_tier_reduction = float(price_tier.reduction_percentage)
                price_tier_info = {
                    'average_price': avg_price,
                    'tier_name': price_tier.name,
                    'price_range': price_tier.price_range_display(),
                    'reduction': price_tier_reduction
                }
            else:
                # No matching price tier found
                price_tier_info = {
                    'average_price': avg_price,
                    'tier_name': 'No Tier Match',
                    'price_range': 'N/A',
                    'reduction': 0,
                    'warning': 'No price tier configured for this price range'
                }
        except Exception as e:
            # Error getting price tier
            price_tier_info = {
                'average_price': avg_price,
                'tier_name': 'Error',
                'price_range': 'N/A',
                'reduction': 0,
                'error': f'Error determining price tier: {str(e)}'
            }

        resolved_condition_details = selected_condition_details.copy() if selected_condition_details else {}

        if condition_assessments is not None:
            # Calculate total from manual assessments (excluding auto-detected categories)
            manual_assessments = {k: v for k, v in condition_assessments.items()
                                if k not in ['brand_category', 'price_tier']}
            manual_assessments_total = sum(manual_assessments.values())

            # Add auto-detected reductions
            layer2_reduction = manual_assessments_total + brand_category_reduction + price_tier_reduction

            # Apply Layer 2 cap
            layer2_reduction = min(layer2_reduction, float(mileage_config.layer2_max_cap))

            # Build condition breakdown
            condition_breakdown = manual_assessments.copy()
            condition_breakdown['brand_category'] = brand_category_reduction
            condition_breakdown['price_tier'] = price_tier_reduction

            selected_options = (
                VehicleConditionCategory.objects.filter(is_active=True)
                .exclude(category_key__in=['brand_category', 'price_tier'])
                .prefetch_related('options')
                .order_by('order')
            )
            for category in selected_options:
                selected_reduction = manual_assessments.get(category.category_key)
                if selected_reduction is None:
                    continue

                option = next(
                    (
                        candidate for candidate in category.options.all()
                        if float(candidate.reduction_percentage) == float(selected_reduction)
                    ),
                    None,
                )
                if option is None:
                    continue

                resolved_condition_details[category.category_key] = {
                    'category_key': category.category_key,
                    'display_name': category.display_name,
                    'option_code': option.option_code,
                    'label': option.label,
                    'display_value': option.display_value or option.label,
                    'reduction_percentage': float(option.reduction_percentage),
                }
        else:
            # Only auto-detected reductions
            layer2_reduction = brand_category_reduction + price_tier_reduction
            condition_breakdown = {
                'brand_category': brand_category_reduction,
                'price_tier': price_tier_reduction
            }

        # Total reduction
        total_reduction = layer1_reduction + layer2_reduction

        # Calculate final adjusted price
        adjusted_price = avg_price * (1 - total_reduction / 100)
        price_savings = avg_price - adjusted_price

        score = _calculate_condition_score(total_reduction)
        grade = _grade_from_score(score)
        comparable_result = get_comparable_listings(
            estimation_data=estimation_data,
            brand=brand,
            model=model,
            variant=variant,
            year=year,
            recommended_price=adjusted_price,
            page=1,
            page_size=10,
        )
        comparables_preview = comparable_result['items']
        comparables_count = comparable_result['total_count']
        market_price_position = _build_market_price_position(
            comparables=comparables_preview,
            price_range=raw_price_range,
            recommended_price=adjusted_price,
            average_price=avg_price,
        )
        confidence_level = _confidence_level(total_data)
        summary = _build_summary(
            score=score,
            grade=grade,
            user_mileage=user_mileage,
            mileage_reduction=layer1_reduction,
            selected_condition_details=resolved_condition_details,
        )

        # Update result with all calculations
        if user_mileage is not None:
            result.update({
                'user_mileage': int(user_mileage),
                'mileage_diff_percent': round(mileage_diff_percent, 1),
            })

        result.update({
            'layer1_reduction': round(layer1_reduction, 1),
            'layer2_reduction': round(layer2_reduction, 1),
            'total_reduction': round(total_reduction, 1),
            'adjusted_price': round(adjusted_price),
            'price_savings': round(price_savings),
            'condition_breakdown': condition_breakdown,
            'selected_condition_details': resolved_condition_details,
            'brand_category_info': brand_category_info,
            'price_tier_info': price_tier_info,
            'config_version': 'simplified_with_auto_brand_and_price_tier_detection',
            'estimated_market_price': round(avg_price),
            'total_reduction_percentage': round(total_reduction, 1),
            'score': score,
            'grade': grade,
            'confidence_level': confidence_level,
            'market_price_position': market_price_position,
            'comparables_count': comparables_count,
            'summary': summary,
        })

        return result
    except Exception as e:
        print(f"Error in get_car_statistics: {e}")
        return None

    return None


def get_client_ip(request):
    """Get client IP address"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def _calculate_condition_score(total_reduction):
    return max(0, round(100 - float(total_reduction)))


def _grade_from_score(score):
    if score >= 90:
        return 'A'
    if score >= 80:
        return 'B'
    if score >= 70:
        return 'C'
    return 'D'


def _confidence_level(sample_size):
    sample_size = int(sample_size or 0)
    if sample_size >= 6:
        return 'high'
    if sample_size >= 3:
        return 'medium'
    return 'low'


def _mileage_impact_label(mileage_reduction):
    reduction = float(mileage_reduction or 0)
    if reduction <= 0:
        return 'Low'
    if reduction < 5:
        return 'Light'
    if reduction < 10:
        return 'Moderate'
    return 'High'


def _condition_label_from_grade(grade):
    return {
        'A': 'Excellent',
        'B': 'Good',
        'C': 'Fair',
        'D': 'Needs Attention',
    }.get(grade, 'Unknown')


def _format_listing_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value in [None, '']:
        return None
    return str(value)


def _safe_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _safe_int(value, default=None):
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _normalize_comparable_from_detail(detail, source, recommended_price):
    if not isinstance(detail, dict):
        return None

    price = _safe_int(detail.get('price'))
    brand = (detail.get('brand') or '').strip()
    model = (detail.get('model') or '').strip()
    variant = (detail.get('variant') or '').strip()
    listing_title = " ".join(part for part in [brand, model, variant] if part).strip()

    comparable = {
        'id': detail.get('id'),
        'listing': {
            'brand': brand,
            'model': model,
            'variant': variant,
            'title': listing_title,
            'location': detail.get('location'),
            'information_ads_date': _format_listing_date(
                detail.get('information_ads_date') or detail.get('created_at')
            ),
        },
        'year': _safe_int(detail.get('year')),
        'mileage': _safe_int(detail.get('mileage')),
        'price': price,
        'appraisal_delta': None if price is None else round(price - float(recommended_price)),
        'source': source or detail.get('source'),
    }
    return comparable


def _normalize_comparable_from_row(row, recommended_price):
    if not isinstance(row, (list, tuple)) or len(row) < 8:
        return None

    price = _safe_int(row[7])
    comparable = {
        'id': row[0],
        'listing': {
            'brand': row[2],
            'model': row[3],
            'variant': row[4],
            'title': " ".join(str(part).strip() for part in [row[2], row[3], row[4]] if part).strip(),
            'location': None,
            'information_ads_date': None,
        },
        'year': _safe_int(row[5]),
        'mileage': _safe_int(row[6]),
        'price': price,
        'appraisal_delta': None if price is None else round(price - float(recommended_price)),
        'source': row[1],
    }
    return comparable


def get_comparable_listings(estimation_data, brand, model, variant, year, recommended_price, page=1, page_size=20):
    page = max(int(page or 1), 1)
    page_size = max(int(page_size or 20), 1)
    comparables = []

    raw_comparables = None
    if isinstance(estimation_data, dict):
        for key in ['comparables', 'comparable_listings', 'listings']:
            if isinstance(estimation_data.get(key), list):
                raw_comparables = estimation_data.get(key)
                break

    if raw_comparables:
        for item in raw_comparables:
            if isinstance(item, dict):
                comparable = _normalize_comparable_from_detail(item, item.get('source'), recommended_price)
                if comparable:
                    comparables.append(comparable)
        total_count = len(comparables)
        start = (page - 1) * page_size
        end = start + page_size
        return {
            'items': comparables[start:end],
            'total_count': total_count,
            'page': page,
            'page_size': page_size,
            'total_pages': ((total_count - 1) // page_size + 1) if total_count else 0,
        }

    start = (page - 1) * page_size
    try:
        records = get_car_records(
            draw=1,
            start=start,
            length=page_size,
            brand_filter=brand,
            model_filter=model,
            variant_filter=variant,
            year_value=year,
        )
    except APIError:
        return {
            'items': [],
            'total_count': 0,
            'page': page,
            'page_size': page_size,
            'total_pages': 0,
        }

    rows = records.get('data', []) if isinstance(records, dict) else []
    total_count = 0
    if isinstance(records, dict):
        total_count = _safe_int(records.get('recordsFiltered'))
        if total_count is None:
            total_count = _safe_int(records.get('recordsTotal'), 0)

    for row in rows:
        comparable = _normalize_comparable_from_row(row, recommended_price)
        if comparable is None:
            continue

        detail = None
        if comparable.get('id') is not None and comparable.get('source'):
            try:
                detail = get_car_detail(comparable['id'], comparable['source'])
            except APIError:
                detail = None

        if isinstance(detail, dict):
            comparable = _normalize_comparable_from_detail(detail, comparable['source'], recommended_price) or comparable

        comparables.append(comparable)

    return {
        'items': comparables,
        'total_count': total_count if total_count is not None else len(comparables),
        'page': page,
        'page_size': page_size,
        'total_pages': ((max(total_count or 0, len(comparables)) - 1) // page_size + 1) if max(total_count or 0, len(comparables)) else 0,
    }


def _build_market_price_position(comparables, price_range, recommended_price, average_price):
    price_range = price_range if isinstance(price_range, dict) else {}
    floor_price = _safe_int(price_range.get('min'))
    if floor_price is None:
        floor_price = _safe_int(price_range.get('low'))

    ceiling_price = _safe_int(price_range.get('max'))
    if ceiling_price is None:
        ceiling_price = _safe_int(price_range.get('high'))

    if floor_price is None or ceiling_price is None:
        prices = [item['price'] for item in comparables if item.get('price') is not None]
        if prices:
            floor_price = min(prices) if floor_price is None else floor_price
            ceiling_price = max(prices) if ceiling_price is None else ceiling_price

    if floor_price is None or ceiling_price is None:
        fallback_prices = [
            floor_price,
            ceiling_price,
            _safe_int(price_range.get('avg')),
            _safe_int(average_price),
        ]
        fallback_prices = [price for price in fallback_prices if price is not None]
        floor_price = min(fallback_prices) if floor_price is None and fallback_prices else floor_price
        ceiling_price = max(fallback_prices) if ceiling_price is None and fallback_prices else ceiling_price

    if floor_price is None:
        floor_price = round(float(recommended_price))
    if ceiling_price is None:
        ceiling_price = round(float(recommended_price))

    return {
        'floor_price': floor_price,
        'recommended_price': round(float(recommended_price)),
        'ceiling_price': ceiling_price,
    }


def _build_summary(score, grade, user_mileage, mileage_reduction, selected_condition_details):
    summary_items = [
        {
            'key': 'condition',
            'label': 'Condition',
            'value': _condition_label_from_grade(grade),
            'display_value': f"{_condition_label_from_grade(grade)} (Grade {grade})",
        },
        {
            'key': 'mileage_score',
            'label': 'Mileage Score',
            'value': user_mileage,
            'display_value': (
                'Mileage not provided'
                if user_mileage in [None, '']
                else f"{_safe_int(user_mileage, 0):,} km - {_mileage_impact_label(mileage_reduction)}"
            ),
        },
    ]

    for category_key in ['accident_history', 'service_history', 'market_demand', 'number_of_owners']:
        detail = selected_condition_details.get(category_key)
        if not detail:
            continue
        summary_items.append({
            'key': category_key,
            'label': detail['display_name'],
            'value': detail['label'],
            'display_value': detail['display_value'],
            'option_code': detail['option_code'],
            'reduction_percentage': detail['reduction_percentage'],
        })

    return {
        'score': score,
        'grade': grade,
        'condition_label': _condition_label_from_grade(grade),
        'items': summary_items,
    }




def normalize_phone_number(phone, country_code):
    """Normalize phone number format"""
    # Remove all non-digits
    phone_digits = re.sub(r'\D', '', phone)

    # Add country code if not present
    if not phone_digits.startswith('60') and not phone_digits.startswith('62'):
        country_digits = country_code.replace('+', '')
        phone_digits = country_digits + phone_digits

    return '+' + phone_digits


def generate_otp():
    """Generate 6-digit OTP for CopyCode"""
    return str(random.randint(100000, 999999))


def is_staff_user(user):
    """Check if user is staff"""
    return user.is_authenticated and user.is_staff


def export_verified_phones(request, export_format):
    """Export verified phones data in CSV or Excel format"""
    try:
        import csv
        from datetime import datetime

        # Build queryset with filters
        queryset = VerifiedPhone.objects.all()

        # Apply filters
        status_filter = request.GET.get('status_filter')
        if status_filter == 'active':
            queryset = queryset.filter(is_active=True)
        elif status_filter == 'inactive':
            queryset = queryset.filter(is_active=False)
        elif status_filter == 'expired':
            queryset = queryset.filter(
                is_active=True,
                verified_at__lt=timezone.now() - timezone.timedelta(days=30)
            )

        search_value = request.GET.get('search[value]', '').strip()
        if search_value:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(phone_number__icontains=search_value) |
                Q(ip_address__icontains=search_value) |
                Q(user_agent__icontains=search_value)
            )

        queryset = queryset.order_by('-id')

        if export_format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="verified_phones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'

            writer = csv.writer(response)
            writer.writerow(['ID', 'Phone Number', 'Verified At', 'Last Accessed', 'Access Count', 'Status', 'IP Address', 'User Agent'])

            for phone in queryset:
                is_expired = phone.is_expired()
                if not phone.is_active:
                    status = 'Inactive'
                elif is_expired:
                    status = 'Expired'
                else:
                    status = 'Active'

                writer.writerow([
                    phone.id,
                    phone.phone_number,
                    phone.verified_at.strftime('%Y-%m-%d %H:%M:%S'),
                    phone.last_accessed.strftime('%Y-%m-%d %H:%M:%S'),
                    phone.access_count,
                    status,
                    phone.ip_address or '',
                    phone.user_agent or ''
                ])

            return response

        elif export_format == 'excel':
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill
                from io import BytesIO

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Verified Phones"

                # Headers
                headers = ['ID', 'Phone Number', 'Verified At', 'Last Accessed', 'Access Count', 'Status', 'IP Address', 'User Agent']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

                # Data
                for row, phone in enumerate(queryset, 2):
                    is_expired = phone.is_expired()
                    if not phone.is_active:
                        status = 'Inactive'
                    elif is_expired:
                        status = 'Expired'
                    else:
                        status = 'Active'

                    ws.cell(row=row, column=1, value=phone.id)
                    ws.cell(row=row, column=2, value=phone.phone_number)
                    ws.cell(row=row, column=3, value=phone.verified_at.strftime('%Y-%m-%d %H:%M:%S'))
                    ws.cell(row=row, column=4, value=phone.last_accessed.strftime('%Y-%m-%d %H:%M:%S'))
                    ws.cell(row=row, column=5, value=phone.access_count)
                    ws.cell(row=row, column=6, value=status)
                    ws.cell(row=row, column=7, value=phone.ip_address or '')
                    ws.cell(row=row, column=8, value=phone.user_agent or '')

                # Auto-size columns
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

                # Save to BytesIO
                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)

                response = HttpResponse(
                    buffer.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="verified_phones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

                return response

            except ImportError:
                # Fallback to CSV if openpyxl is not available
                return export_verified_phones(request, 'csv')

    except Exception as e:
        from django.http import JsonResponse
        return JsonResponse({'error': f'Export failed: {str(e)}'}, status=500)


def export_otp_sessions(request, export_format):
    """Export OTP sessions data in CSV or Excel format"""
    try:
        import csv
        from datetime import datetime
        from ..models import OTPSession

        # Build queryset with filters
        queryset = OTPSession.objects.all()

        # Apply filters
        status_filter = request.GET.get('status_filter')
        if status_filter == 'used':
            queryset = queryset.filter(is_used=True)
        elif status_filter == 'unused':
            queryset = queryset.filter(is_used=False)
        elif status_filter == 'expired':
            queryset = queryset.filter(
                is_used=False,
                created_at__lt=timezone.now() - timezone.timedelta(minutes=1)
            )

        search_value = request.GET.get('search[value]', '').strip()
        if search_value:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(phone_number__icontains=search_value) |
                Q(otp_code__icontains=search_value) |
                Q(ip_address__icontains=search_value)
            )

        queryset = queryset.order_by('-id')

        if export_format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="otp_sessions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'

            writer = csv.writer(response)
            writer.writerow(['ID', 'Phone Number', 'OTP Code', 'Created At', 'Status', 'IP Address'])

            for otp in queryset:
                is_expired = otp.is_expired()

                if otp.is_used:
                    status = 'Used'
                elif is_expired:
                    status = 'Expired'
                else:
                    status = 'Not Used'

                # Mask phone number
                masked_phone = otp.phone_number[:4] + '*' * (len(otp.phone_number) - 8) + otp.phone_number[-4:] if otp.phone_number else ''

                # Mask OTP code for security
                masked_otp = otp.otp_code[:2] + '****' if otp.otp_code else 'N/A'

                writer.writerow([
                    otp.id,
                    masked_phone,
                    masked_otp,
                    otp.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    status,
                    otp.ip_address or ''
                ])

            return response

        elif export_format == 'excel':
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill
                from io import BytesIO

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "OTP Sessions"

                # Headers
                headers = ['ID', 'Phone Number', 'OTP Code', 'Created At', 'Status', 'IP Address']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

                # Data
                for row, otp in enumerate(queryset, 2):
                    is_expired = otp.is_expired()

                    if otp.is_used:
                        status = 'Used'
                    elif is_expired:
                        status = 'Expired'
                    else:
                        status = 'Not Used'

                    # Mask phone number and OTP code
                    masked_phone = otp.phone_number[:4] + '*' * (len(otp.phone_number) - 8) + otp.phone_number[-4:] if otp.phone_number else ''
                    masked_otp = otp.otp_code[:2] + '****' if otp.otp_code else 'N/A'

                    ws.cell(row=row, column=1, value=otp.id)
                    ws.cell(row=row, column=2, value=masked_phone)
                    ws.cell(row=row, column=3, value=masked_otp)
                    ws.cell(row=row, column=4, value=otp.created_at.strftime('%Y-%m-%d %H:%M:%S'))
                    ws.cell(row=row, column=5, value=status)
                    ws.cell(row=row, column=6, value=otp.ip_address or '')

                # Auto-size columns
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

                # Save to BytesIO
                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)

                response = HttpResponse(
                    buffer.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="otp_sessions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

                return response

            except ImportError:
                # Fallback to CSV if openpyxl is not available
                return export_otp_sessions(request, 'csv')

    except Exception as e:
        from django.http import JsonResponse
        return JsonResponse({'error': f'Export failed: {str(e)}'}, status=500)

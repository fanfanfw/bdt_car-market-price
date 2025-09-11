from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.db import connection, models
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from datetime import timedelta
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.views import LoginView
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
import json
import random
import re
import requests
import os
from django.conf import settings
from decouple import config
from .models import (
    CarStandard, CarUnified, Category, BrandCategory, VerifiedPhone, OTPSession,
    MileageConfiguration, VehicleConditionCategory, ConditionOption, PriceTier
)

def index(request):
    """Main index page with car price estimation form"""
    # Get all active vehicle condition categories with their options
    # Exclude brand_category and price_tier as they will be handled automatically
    categories = VehicleConditionCategory.objects.filter(
        is_active=True
    ).exclude(category_key__in=['brand_category', 'price_tier']).prefetch_related('options').order_by('order')
    
    context = {
        'condition_categories': categories,
    }
    return render(request, 'main/index.html', context)

def result(request):
    """Display calculation result page with secure OTP verification"""
    context = {}
    
    if request.method == 'POST':
        # Handle form submission directly
        brand = request.POST.get('brand')
        model = request.POST.get('model')
        variant = request.POST.get('variant')
        year = request.POST.get('year')
        user_mileage = request.POST.get('user_mileage')
        
        # Get condition assessment values (excluding auto-detected categories)
        condition_assessments = {
            'exterior_condition': float(request.POST.get('exterior_condition', 0)),
            'interior_condition': float(request.POST.get('interior_condition', 0)),
            'mechanical_condition': float(request.POST.get('mechanical_condition', 0)),
            'accident_history': float(request.POST.get('accident_history', 0)),
            'service_history': float(request.POST.get('service_history', 0)),
            'number_of_owners': float(request.POST.get('number_of_owners', 0)),
            'tires_brakes': float(request.POST.get('tires_brakes', 0)),
            'modifications': float(request.POST.get('modifications', 0)),
            'market_demand': float(request.POST.get('market_demand', 0)),
            # brand_category and price_tier are now auto-detected, not from form
        }
        
        if brand and model and variant and year:
            # Store form data in session for security (encrypted)
            request.session['calculation_request'] = {
                'brand': brand,
                'model': model,
                'variant': variant,
                'year': int(year),
                'user_mileage': user_mileage,
                'condition_assessments': condition_assessments
            }
            
            # Check if user has verified phone in cookie (1 day session)
            verified_phone_cookie = request.COOKIES.get('verified_phone')
            phone_already_verified = False
            cookie_should_be_deleted = False
            
            if verified_phone_cookie:
                # Check if phone is still active in database (1 month verification)
                try:
                    verified_phone = VerifiedPhone.objects.get(phone_number=verified_phone_cookie)
                    if not verified_phone.is_expired() and verified_phone.is_active:
                        phone_already_verified = True
                        context['verified_phone'] = verified_phone_cookie
                    else:
                        # Phone expired, mark for cookie deletion
                        cookie_should_be_deleted = True
                except VerifiedPhone.DoesNotExist:
                    # Phone not found in database, mark for cookie deletion
                    cookie_should_be_deleted = True
            
            context['phone_not_verified'] = not phone_already_verified
            context['car_info'] = f"{brand} {model} {variant} ({year})"
            
            # If phone already verified, we can show results immediately via JavaScript
            if phone_already_verified:
                context['skip_otp'] = True
        else:
            messages.error(request, 'Please complete all required data.')
            return redirect('main:index')
    else:
        # GET request - redirect to index
        messages.info(request, 'Please fill out the form first.')
        return redirect('main:index')
    
    # Create response and handle cookie deletion if needed
    response = render(request, 'main/result.html', context)
    
    if cookie_should_be_deleted:
        response.delete_cookie('verified_phone')
    
    return response

def get_categories(request):
    """API endpoint to get all categories"""
    try:
        categories = Category.objects.values_list('name', flat=True).order_by('name')
        return JsonResponse(list(categories), safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_brands(request):
    """API endpoint to get all brands"""
    try:
        # Get all brands directly from CarStandard
        brands = CarStandard.objects.values_list('brand_norm', flat=True).distinct().order_by('brand_norm')
        return JsonResponse(list(brands), safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_models(request):
    """API endpoint to get models for selected brand"""
    try:
        brand = request.GET.get('brand')
        if not brand:
            return JsonResponse({'error': 'Brand parameter required'}, status=400)
        
        models = CarStandard.objects.filter(
            brand_norm=brand
        ).values_list('model_norm', flat=True).distinct().order_by('model_norm')
        
        return JsonResponse(list(models), safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_variants(request):
    """API endpoint to get variants for selected brand and model"""
    try:
        brand = request.GET.get('brand')
        model = request.GET.get('model')
        
        if not brand or not model:
            return JsonResponse({'error': 'Brand and model parameters required'}, status=400)
        
        variants = CarStandard.objects.filter(
            brand_norm=brand,
            model_norm=model
        ).values_list('variant_norm', flat=True).distinct().order_by('variant_norm')
        
        return JsonResponse(list(variants), safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_years(request):
    """API endpoint to get years for selected brand, model, and variant"""
    try:
        brand = request.GET.get('brand')
        model = request.GET.get('model')
        variant = request.GET.get('variant')
        
        if not brand or not model or not variant:
            return JsonResponse({'error': 'Brand, model, and variant parameters required'}, status=400)
        
        # Get cars_standard_id first
        cars_standard = CarStandard.objects.filter(
            brand_norm=brand,
            model_norm=model,
            variant_norm=variant
        ).first()
        
        if not cars_standard:
            return JsonResponse([], safe=False)
        
        # Get years from cars_unified table
        years = CarUnified.objects.filter(
            cars_standard=cars_standard,
            year__isnull=False
        ).values_list('year', flat=True).distinct().order_by('-year')
        
        return JsonResponse(list(years), safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_mileage_config():
    """Get the mileage configuration"""
    try:
        return MileageConfiguration.objects.first()
    except MileageConfiguration.DoesNotExist:
        # Return default values if no config exists
        return type('obj', (object,), {
            'threshold_percent': 10.0,
            'reduction_percent': 2.0,
            'max_reduction_cap': 15.0
        })

def get_car_statistics(brand, model, variant, year, user_mileage=None, condition_assessments=None):
    """Get car statistics including average mileage and price with condition assessments"""
    try:
        # Get mileage configuration
        mileage_config = get_mileage_config()

        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT
                  s.brand_norm,
                  s.model_norm,
                  s.variant_norm,
                  c.year,
                  ROUND(AVG(c.mileage)) AS rata_rata_mileage_bulat,
                  ROUND(AVG(c.price)) AS rata_rata_price_bulat,
                  COUNT(*) AS total_data
                FROM
                  public.cars_unified c
                JOIN
                  public.cars_standard s
                ON
                  c.cars_standard_id = s.id
                WHERE
                  s.brand_norm = %s
                  AND s.model_norm = %s
                  AND s.variant_norm = %s
                  AND c.year = %s
                  AND c.mileage IS NOT NULL
                  AND c.price IS NOT NULL
                GROUP BY
                  s.brand_norm,
                  s.model_norm,
                  s.variant_norm,
                  c.year
            """, [brand, model, variant, year])
            
            row = cursor.fetchone()
            if row:
                result = {
                    'brand_norm': row[0],
                    'model_norm': row[1],
                    'variant_norm': row[2],
                    'year': row[3],
                    'rata_rata_mileage_bulat': row[4],
                    'rata_rata_price_bulat': row[5],
                    'total_data': row[6],
                }
                
                # Calculate price adjustments with dynamic 2-layer system
                avg_mileage = float(row[4])
                avg_price = float(row[5])
                
                # Layer 1: Mileage-based reduction
                layer1_reduction = 0
                mileage_diff_percent = 0
                
                if user_mileage is not None:
                    user_mileage = float(user_mileage)
                    if user_mileage > avg_mileage:
                        mileage_diff_percent = ((user_mileage - avg_mileage) / avg_mileage) * 100
                        threshold = float(mileage_config.threshold_percent)
                        reduction_per_threshold = float(mileage_config.reduction_percent)
                        layer1_reduction = (mileage_diff_percent / threshold) * reduction_per_threshold
                        # Apply cap
                        layer1_reduction = min(layer1_reduction, float(mileage_config.max_reduction_cap))
                    else:
                        mileage_diff_percent = ((user_mileage - avg_mileage) / avg_mileage) * 100
                
                # Layer 2: Condition assessment reduction + Brand Category Auto-Detection
                layer2_reduction = 0
                condition_breakdown = {}
                brand_category_reduction = 0
                brand_category_info = None
                
                # Auto-detect brand category reduction
                try:
                    brand_category_mapping = BrandCategory.objects.select_related('category').get(brand=brand)
                    # Get the reduction percentage directly from the category model
                    brand_category_reduction = float(brand_category_mapping.category.reduction_percentage)
                    brand_category_info = {
                        'brand': brand,
                        'category': brand_category_mapping.category.name,
                        'reduction': brand_category_reduction
                    }
                except BrandCategory.DoesNotExist:
                    # Brand not classified - use 0% reduction
                    brand_category_reduction = 0
                    brand_category_info = {
                        'brand': brand,
                        'category': 'Unclassified',
                        'reduction': 0,
                        'warning': 'Brand not classified - admin should classify this brand'
                    }
                
                # Auto-detect price tier reduction
                price_tier_reduction = 0
                price_tier_info = None
                
                try:
                    price_tier = PriceTier.get_tier_for_price(avg_price)
                    if price_tier:
                        price_tier_reduction = float(price_tier.reduction_percentage)
                        price_tier_info = {
                            'average_price': avg_price,
                            'tier_name': price_tier.name,
                            'price_range': price_tier.price_range_display(),
                            'reduction': price_tier_reduction
                        }
                    else:
                        # No matching price tier found
                        price_tier_info = {
                            'average_price': avg_price,
                            'tier_name': 'No Tier Match',
                            'price_range': 'N/A',
                            'reduction': 0,
                            'warning': 'No price tier configured for this price range'
                        }
                except Exception as e:
                    # Error getting price tier
                    price_tier_info = {
                        'average_price': avg_price,
                        'tier_name': 'Error',
                        'price_range': 'N/A',
                        'reduction': 0,
                        'error': f'Error determining price tier: {str(e)}'
                    }
                
                if condition_assessments is not None:
                    # Calculate total from manual assessments (excluding auto-detected categories)
                    manual_assessments = {k: v for k, v in condition_assessments.items() 
                                        if k not in ['brand_category', 'price_tier']}
                    manual_assessments_total = sum(manual_assessments.values())
                    
                    # Add auto-detected reductions
                    layer2_reduction = manual_assessments_total + brand_category_reduction + price_tier_reduction
                    
                    # Apply Layer 2 cap
                    layer2_reduction = min(layer2_reduction, float(mileage_config.layer2_max_cap))
                    
                    # Build condition breakdown
                    condition_breakdown = manual_assessments.copy()
                    condition_breakdown['brand_category'] = brand_category_reduction
                    condition_breakdown['price_tier'] = price_tier_reduction
                else:
                    # Only auto-detected reductions
                    layer2_reduction = brand_category_reduction + price_tier_reduction
                    condition_breakdown = {
                        'brand_category': brand_category_reduction,
                        'price_tier': price_tier_reduction
                    }
                
                # Total reduction
                total_reduction = layer1_reduction + layer2_reduction
                
                # Calculate final adjusted price
                adjusted_price = avg_price * (1 - total_reduction / 100)
                price_savings = avg_price - adjusted_price
                
                # Update result with all calculations
                if user_mileage is not None:
                    result.update({
                        'user_mileage': int(user_mileage),
                        'mileage_diff_percent': round(mileage_diff_percent, 1),
                    })
                
                result.update({
                    'layer1_reduction': round(layer1_reduction, 1),
                    'layer2_reduction': round(layer2_reduction, 1),
                    'total_reduction': round(total_reduction, 1),
                    'adjusted_price': round(adjusted_price),
                    'price_savings': round(price_savings),
                    'condition_breakdown': condition_breakdown,
                    'brand_category_info': brand_category_info,
                    'price_tier_info': price_tier_info,
                    'config_version': 'simplified_with_auto_brand_and_price_tier_detection'
                })
                
                return result
    except Exception as e:
        print(f"Error in get_car_statistics: {e}")
        return None
    
    return None

def get_client_ip(request):
    """Get client IP address"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def format_phone_number_for_message_central(phone, country_code):
    """Format phone number for Message_Central API (Indonesia +62 and Malaysia +60)"""
    # Remove all non-digits from phone
    cleaned_phone = re.sub(r'[^\d]', '', phone)
    # Extract country code number (remove + sign)
    country_num = re.sub(r'[^\d]', '', country_code)
    
    # Handle country code formatting according to Message_Central API
    if country_num == '62':
        # Indonesia: countryCode=62, mobileNumber=89525521887
        country_code_formatted = '62'
        mobile_number = cleaned_phone
        
        # Remove leading 0 if present (08xxx -> 8xxx)
        if mobile_number.startswith('0'):
            mobile_number = mobile_number[1:]
            
        return country_code_formatted, mobile_number
        
    elif country_num == '60':
        # Malaysia: countryCode=60, mobileNumber=173023419
        country_code_formatted = '60'
        mobile_number = cleaned_phone
        
        # Remove leading 0 if present (01xxx -> 1xxx)
        if mobile_number.startswith('0'):
            mobile_number = mobile_number[1:]
            
        return country_code_formatted, mobile_number
        
    else:
        # Unsupported country
        return None, None

def normalize_phone_number(phone, country_code):
    """Normalize phone number format (legacy function for backward compatibility)"""
    # Remove all non-digits
    phone_digits = re.sub(r'\D', '', phone)
    
    # Add country code if not present
    if not phone_digits.startswith('60') and not phone_digits.startswith('62'):
        country_digits = country_code.replace('+', '')
        phone_digits = country_digits + phone_digits
    
    return '+' + phone_digits

def generate_otp():
    """Generate 6-digit OTP"""
    return str(random.randint(100000, 999999))

@csrf_exempt
@require_http_methods(["POST"])
def check_phone_status(request):
    """Check if phone number is already verified and active"""
    try:
        data = json.loads(request.body)
        phone = data.get('phone')
        country_code = data.get('country_code', '+60')
        
        if not phone:
            return JsonResponse({'error': 'Phone number required'}, status=400)
        
        # Normalize phone number
        full_phone = normalize_phone_number(phone, country_code)
        
        # Check if phone exists and is active
        try:
            verified_phone = VerifiedPhone.objects.get(phone_number=full_phone)
            
            if verified_phone.is_expired():
                # Phone expired, mark as inactive
                verified_phone.is_active = False
                verified_phone.save()
                return JsonResponse({
                    'verified': False,
                    'expired': True,
                    'message': 'Phone verification has expired. Please verify again.'
                })
            
            # Phone is active and valid
            verified_phone.access_count += 1
            verified_phone.save()
            
            # Set cookie for user convenience (same as OTP verification)
            response = JsonResponse({
                'verified': True,
                'phone': full_phone,
                'message': 'Phone number is already verified.'
            })
            
            cookie_age = getattr(settings, 'OTP_SESSION_COOKIE_AGE', 86400)
            response.set_cookie(
                'verified_phone',
                full_phone,
                max_age=cookie_age,
                httponly=False,  # Allow JavaScript access
                samesite='Strict'
            )
            
            return response
            
        except VerifiedPhone.DoesNotExist:
            return JsonResponse({
                'verified': False,
                'message': 'Phone number not verified yet.'
            })
            
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
def send_otp(request):
    """Send OTP to phone number using Message_Central API"""
    try:
        data = json.loads(request.body)
        phone = data.get('phone')
        country_code = data.get('country_code', '+60')
        
        if not phone:
            return JsonResponse({'error': 'Phone number required'}, status=400)
        
        # Format phone number for Message_Central API
        country_code_formatted, mobile_number = format_phone_number_for_message_central(phone, country_code)
        
        if not country_code_formatted or not mobile_number:
            return JsonResponse({'error': 'Invalid phone number format. Use +62 for Indonesia or +60 for Malaysia'}, status=400)
        
        # Create normalized full phone number for our database
        full_phone = normalize_phone_number(phone, country_code)
        
        # Validate phone number format
        if country_code == '+60':
            # Malaysia format validation
            if not re.match(r'^\+60\d{8,10}$', full_phone):
                return JsonResponse({'error': 'Invalid Malaysian phone number format'}, status=400)
        elif country_code == '+62':
            # Indonesia format validation
            if not re.match(r'^\+62\d{8,12}$', full_phone):
                return JsonResponse({'error': 'Invalid Indonesian phone number format'}, status=400)
        
        # Get Message_Central API configuration from environment
        auth_token = config('MESSAGE_CENTRAL_AUTH_TOKEN', default=None)
        customer_id = config('MESSAGE_CENTRAL_CUSTOMER_ID', default=None)
        
        if not auth_token or not customer_id:
            return JsonResponse({'error': 'Message_Central API configuration not found'}, status=500)
        
        # Call Message_Central API to send OTP
        try:
            url = f"https://cpaas.messagecentral.com/verification/v3/send?countryCode={country_code_formatted}&customerId={customer_id}&flowType=WHATSAPP&mobileNumber={mobile_number}"
            
            headers = {
                'authToken': auth_token
            }
            
            response = requests.post(url, headers=headers, data={}, timeout=30)
            response_data = response.json()
            
            if response.status_code == 200 and response_data.get('responseCode') == 200:
                # Success - get verificationId from response
                verification_id = response_data.get('data', {}).get('verificationId')
                
                if not verification_id:
                    return JsonResponse({'error': 'No verification ID received from Message_Central'}, status=500)
                
                # Clean up old OTP sessions for this phone
                OTPSession.objects.filter(phone_number=full_phone, is_used=False).update(is_used=True)
                
                # Create new OTP session with verificationId
                otp_session = OTPSession.objects.create(
                    phone_number=full_phone,
                    verification_id=verification_id,
                    ip_address=get_client_ip(request)
                )
                
                return JsonResponse({
                    'success': True,
                    'message': f'OTP sent to {full_phone} via WhatsApp',
                    'verification_id': verification_id,
                    'expires_in': 60  # seconds
                })
            else:
                # Error from Message_Central
                error_message = response_data.get('message', 'Failed to send OTP')
                return JsonResponse({'error': f'Message_Central Error: {error_message}'}, status=400)
                
        except requests.exceptions.Timeout:
            return JsonResponse({'error': 'Request timeout. Please try again.'}, status=500)
        except requests.exceptions.RequestException as e:
            return JsonResponse({'error': f'Network error: {str(e)}'}, status=500)
        except Exception as e:
            return JsonResponse({'error': f'API error: {str(e)}'}, status=500)
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
def verify_otp(request):
    """Verify OTP using Message_Central API and mark phone as verified"""
    try:
        data = json.loads(request.body)
        phone = data.get('phone')
        otp_code = data.get('otp')
        country_code = data.get('country_code', '+60')
        
        if not phone or not otp_code:
            return JsonResponse({'error': 'Phone number and OTP required'}, status=400)
        
        # Validate OTP code format (should be 4 digits)
        if not otp_code or not otp_code.isdigit() or len(otp_code) != 4:
            return JsonResponse({'error': 'OTP code must be 4 digits'}, status=400)
        
        # Format phone number for Message_Central API
        country_code_formatted, mobile_number = format_phone_number_for_message_central(phone, country_code)
        
        if not country_code_formatted or not mobile_number:
            return JsonResponse({'error': 'Invalid phone number format. Use +62 for Indonesia or +60 for Malaysia'}, status=400)
        
        # Create normalized full phone number for our database
        full_phone = normalize_phone_number(phone, country_code)
        
        # Find the OTP session with verificationId
        try:
            otp_session = OTPSession.objects.filter(
                phone_number=full_phone,
                is_used=False
            ).order_by('-created_at').first()
            
            if not otp_session:
                return JsonResponse({'error': 'No active OTP session found. Please request a new OTP.'}, status=400)
            
            if otp_session.is_expired():
                return JsonResponse({'error': 'OTP has expired. Please request a new one.'}, status=400)
            
            # Get verificationId from our session 
            verification_id = otp_session.verification_id
            
            # Get Message_Central API configuration from environment
            auth_token = config('MESSAGE_CENTRAL_AUTH_TOKEN', default=None)
            customer_id = config('MESSAGE_CENTRAL_CUSTOMER_ID', default=None)
            
            if not auth_token or not customer_id:
                return JsonResponse({'error': 'Message_Central API configuration not found'}, status=500)
            
            # Call Message_Central API to validate OTP
            try:
                url = f"https://cpaas.messagecentral.com/verification/v3/validateOtp?countryCode={country_code_formatted}&mobileNumber={mobile_number}&verificationId={verification_id}&customerId={customer_id}&code={otp_code}"
                
                headers = {
                    'authToken': auth_token
                }
                
                response = requests.get(url, headers=headers, timeout=30)
                response_data = response.json()
                
                if response.status_code == 200 and response_data.get('responseCode') == 200:
                    # Check verification status
                    verification_status = response_data.get('data', {}).get('verificationStatus')
                    
                    if verification_status == 'VERIFICATION_COMPLETED':
                        # OTP verification successful
                        
                        # Mark OTP as used
                        otp_session.is_used = True
                        otp_session.save()
                        
                        # Create or update verified phone
                        verified_phone, created = VerifiedPhone.objects.get_or_create(
                            phone_number=full_phone,
                            defaults={
                                'user_agent': request.META.get('HTTP_USER_AGENT', ''),
                                'ip_address': get_client_ip(request),
                                'access_count': 1,
                                'is_active': True
                            }
                        )
                        
                        if not created:
                            # Phone already exists, extend expiry
                            verified_phone.extend_expiry()
                            verified_phone.user_agent = request.META.get('HTTP_USER_AGENT', '')
                            verified_phone.ip_address = get_client_ip(request)
                            verified_phone.access_count += 1
                            verified_phone.save()
                        
                        # Set session cookie for user convenience
                        response = JsonResponse({
                            'success': True,
                            'phone': full_phone,
                            'message': 'Phone number verified successfully!'
                        })
                        
                        cookie_age = getattr(settings, 'OTP_SESSION_COOKIE_AGE', 86400)
                        response.set_cookie(
                            'verified_phone', 
                            full_phone, 
                            max_age=cookie_age,
                            httponly=False,  # Allow JavaScript access for form pre-fill
                            samesite='Strict'
                        )
                        
                        return response
                    else:
                        # OTP verification failed
                        return JsonResponse({'error': 'Invalid OTP code. Please try again.'}, status=400)
                else:
                    # Error from Message_Central
                    error_message = response_data.get('message', 'OTP verification failed')
                    return JsonResponse({'error': f'Verification Error: {error_message}'}, status=400)
                    
            except requests.exceptions.Timeout:
                return JsonResponse({'error': 'Request timeout. Please try again.'}, status=500)
            except requests.exceptions.RequestException as e:
                return JsonResponse({'error': f'Network error: {str(e)}'}, status=500)
            except Exception as e:
                return JsonResponse({'error': f'API error: {str(e)}'}, status=500)
            
        except Exception as e:
            return JsonResponse({'error': 'Verification failed'}, status=500)
            
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["POST"])
def get_secure_results(request):
    """Get calculation results only if phone is verified"""
    try:
        data = json.loads(request.body)
        phone_number = data.get('phone_number')
        
        if not phone_number:
            return JsonResponse({'error': 'Phone number required'}, status=400)
        
        # Verify phone is active
        try:
            verified_phone = VerifiedPhone.objects.get(phone_number=phone_number)
            if verified_phone.is_expired():
                return JsonResponse({'error': 'Phone verification expired'}, status=403)
        except VerifiedPhone.DoesNotExist:
            return JsonResponse({'error': 'Phone not verified'}, status=403)
        
        # Get calculation data from session
        calculation_data = request.session.get('calculation_request')
        if not calculation_data:
            return JsonResponse({'error': 'No calculation data found'}, status=400)
        
        # Perform calculation
        result_data = get_car_statistics(
            calculation_data['brand'],
            calculation_data['model'], 
            calculation_data['variant'],
            calculation_data['year'],
            calculation_data.get('user_mileage'),
            calculation_data.get('condition_assessments')
        )
        
        if result_data:
            # Update access count
            verified_phone.access_count += 1
            verified_phone.save()
            
            # Don't clear session data so user can recalculate with same data
            # Session data will be cleared when user starts new calculation
            
            return JsonResponse({
                'success': True,
                'result': result_data
            })
        else:
            return JsonResponse({
                'success': False,
                'no_data': True,
                'message': 'No data found for the selected combination'
            })
            
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# Admin Views
def is_staff_user(user):
    """Check if user is staff"""
    return user.is_authenticated and user.is_staff

class CustomAdminLoginView(LoginView):
    """Custom admin login view"""
    template_name = 'admin/admin-login.html'
    redirect_authenticated_user = True
    
    def get_success_url(self):
        return '/panel/dashboard/'
    
    def form_valid(self, form):
        """Check if user is staff before allowing login"""
        user = form.get_user()
        
        if not user.is_staff:
            form.add_error(None, 'You do not have admin privileges.')
            return self.form_invalid(form)
        
        messages.success(self.request, f'Welcome back, {user.first_name or user.username}!')
        return super().form_valid(form)
    
    def dispatch(self, request, *args, **kwargs):
        """Redirect if already authenticated and is staff"""
        if request.user.is_authenticated and request.user.is_staff:
            return redirect(self.get_success_url())
        return super().dispatch(request, *args, **kwargs)

@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def admin_dashboard_view(request):
    """Admin dashboard"""
    # Get statistics
    stats = {
        'total_users': User.objects.count(),
        'verified_phones': VerifiedPhone.objects.filter(is_active=True).count(),
        'car_records': CarUnified.objects.count(),
        'today_calculations': 0,  # This would need to be tracked separately
    }
    
    # Get recent activity (mock data for now)
    recent_activities = [
        {
            'icon': 'phone',
            'description': 'New phone verification',
            'timestamp': timezone.now()
        },
        {
            'icon': 'calculator',
            'description': 'Price calculation performed',
            'timestamp': timezone.now()
        },
    ]
    
    context = {
        'stats': stats,
        'recent_activities': recent_activities,
    }
    
    return render(request, 'admin/dashboard.html', context)

@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def admin_logout_view(request):
    """Admin logout"""
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('admin_login')

@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def formula_config_edit(request):
    """Edit formula configuration"""
    config, created = MileageConfiguration.objects.get_or_create(defaults={
        'threshold_percent': 10.0,
        'reduction_percent': 2.0,
        'max_reduction_cap': 15.0,
        'layer2_max_cap': 70.0
    })
    
    if request.method == 'POST':
        try:
            config.threshold_percent = float(request.POST.get('threshold_percent', 10.0))
            config.reduction_percent = float(request.POST.get('reduction_percent', 2.0))
            config.max_reduction_cap = float(request.POST.get('max_reduction_cap', 15.0))
            config.layer2_max_cap = float(request.POST.get('layer2_max_cap', 70.0))
            config.save()
            
            messages.success(request, 'Formula configuration updated successfully.')
            return redirect('main:formula_config_edit')
        except Exception as e:
            messages.error(request, f'Error updating configuration: {str(e)}')
    
    context = {
        'config': config
    }
    return render(request, 'admin/formula-config.html', context)

@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def condition_categories_manage(request):
    """Manage condition categories and their options"""
    # Exclude brand_category and price_tier as they are now handled automatically
    categories = VehicleConditionCategory.objects.prefetch_related('options').exclude(category_key__in=['brand_category', 'price_tier']).order_by('order')
    
    context = {
        'categories': categories
    }
    return render(request, 'admin/condition-categories.html', context)

@csrf_exempt
@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def condition_option_edit(request, option_id):
    """Edit a condition option"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=400)
    
    try:
        option = get_object_or_404(ConditionOption, id=option_id)
        data = json.loads(request.body)
        
        option.label = data.get('label', '').strip()
        option.reduction_percentage = float(data.get('reduction_percentage', 0))
        
        if not option.label:
            return JsonResponse({'error': 'Option label is required'}, status=400)
        
        option.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Option "{option.label}" updated successfully'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def condition_option_add(request, category_id):
    """Add new option to a category"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=400)
    
    try:
        category = get_object_or_404(VehicleConditionCategory, id=category_id)
        data = json.loads(request.body)
        
        label = data.get('label', '').strip()
        reduction_percentage = float(data.get('reduction_percentage', 0))
        
        if not label:
            return JsonResponse({'error': 'Option label is required'}, status=400)
        
        # Check for duplicate labels
        if category.options.filter(label=label).exists():
            return JsonResponse({'error': 'Option with this label already exists'}, status=400)
        
        # Get next order
        next_order = category.options.count()
        
        option = ConditionOption.objects.create(
            category=category,
            label=label,
            reduction_percentage=reduction_percentage,
            order=next_order
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Option "{option.label}" added successfully',
            'option_id': option.id
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def condition_option_delete(request, option_id):
    """Delete a condition option"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=400)
    
    try:
        option = get_object_or_404(ConditionOption, id=option_id)
        category = option.category
        
        # Prevent deleting if only one option left
        if category.options.count() <= 1:
            return JsonResponse({
                'error': 'Cannot delete the last option. At least one option is required.'
            }, status=400)
        
        option_label = option.label
        option.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Option "{option_label}" deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def car_data_view(request):
    """Car database management view with DataTables"""
    context = {
        'page_title': 'Car Database',
        'total_cars': CarUnified.objects.count(),
        'total_brands': CarUnified.objects.values('brand').distinct().count(),
        'total_models': CarUnified.objects.values('model').distinct().count(),
        'sources': CarUnified.SOURCE_CHOICES,
    }
    return render(request, 'admin/car-data.html', context)

@csrf_exempt
@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def car_data_api(request):
    """API endpoint for DataTables car data"""
    try:
        # DataTables parameters
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '').strip()
        
        # Ordering
        order_column_index = int(request.GET.get('order[0][column]', 0))
        order_direction = request.GET.get('order[0][dir]', 'asc')
        
        # Column mapping for ordering
        columns = ['id', 'source', 'brand', 'model', 'variant', 'year', 'mileage', 'price']
        order_column = columns[order_column_index] if order_column_index < len(columns) else 'id'
        
        if order_direction == 'desc':
            order_column = '-' + order_column
        
        # Base queryset
        queryset = CarUnified.objects.all()
        
        # Additional filtering based on request parameters
        source_filter = request.GET.get('source_filter')
        year_filter = request.GET.get('year_filter')
        price_filter = request.GET.get('price_filter')
        
        if source_filter:
            queryset = queryset.filter(source=source_filter)
            
        if year_filter:
            if year_filter == '2024-':
                queryset = queryset.filter(year__gte=2024)
            elif year_filter == '2020-2023':
                queryset = queryset.filter(year__gte=2020, year__lte=2023)
            elif year_filter == '2015-2019':
                queryset = queryset.filter(year__gte=2015, year__lte=2019)
            elif year_filter == '2010-2014':
                queryset = queryset.filter(year__gte=2010, year__lte=2014)
            elif year_filter == '-2009':
                queryset = queryset.filter(year__lte=2009)
                
        if price_filter:
            if price_filter == '0-50000':
                queryset = queryset.filter(price__gte=0, price__lte=50000)
            elif price_filter == '50000-100000':
                queryset = queryset.filter(price__gte=50000, price__lte=100000)
            elif price_filter == '100000-200000':
                queryset = queryset.filter(price__gte=100000, price__lte=200000)
            elif price_filter == '200000-':
                queryset = queryset.filter(price__gte=200000)
        
        # Search filtering
        if search_value:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(brand__icontains=search_value) |
                Q(model__icontains=search_value) |
                Q(variant__icontains=search_value) |
                Q(condition__icontains=search_value) |
                Q(location__icontains=search_value) |
                Q(source__icontains=search_value)
            )
        
        # Total records
        total_records = CarUnified.objects.count()
        filtered_records = queryset.count()
        
        # Apply ordering and pagination
        queryset = queryset.order_by(order_column)[start:start + length]
        
        # Build data for DataTables
        data = []
        for car in queryset:
            # Format price
            price_formatted = f"RM {car.price:,}" if car.price else "-"
            
            # Format mileage
            mileage_formatted = f"{car.mileage:,} km" if car.mileage else "-"
            
            # Format year
            year_formatted = str(car.year) if car.year else "-"
            
            # Format condition
            condition_formatted = car.condition.title() if car.condition else "-"
            
            # Format source
            source_formatted = dict(CarUnified.SOURCE_CHOICES).get(car.source, car.source)
            
            # Format location (truncate if too long)
            location_formatted = (car.location[:30] + '...') if car.location and len(car.location) > 30 else (car.location or "-")
            
            # Actions column
            actions = f'''<div class="btn-group btn-group-sm" role="group">
                <a href="{car.listing_url}" target="_blank" class="btn btn-info btn-sm" title="View Listing">
                    <i class="fas fa-external-link-alt"></i>
                </a>
                <button type="button" class="btn btn-primary btn-sm" onclick="viewCarDetails({car.id})" title="View Details">
                    <i class="fas fa-eye"></i>
                </button>
            </div>'''
            
            data.append([
                car.id,
                source_formatted,
                car.brand,
                car.model,
                car.variant or "-",
                year_formatted,
                car.mileage,  # Raw mileage data
                car.price,    # Raw price data
                actions
            ])
        
        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': filtered_records,
            'data': data
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def car_detail_api(request, car_id):
    """API endpoint to get detailed car information"""
    try:
        car = get_object_or_404(CarUnified, id=car_id)
        
        # Get related standard car info if available
        standard_info = None
        if car.cars_standard:
            standard_info = {
                'brand_norm': car.cars_standard.brand_norm,
                'model_norm': car.cars_standard.model_norm,
                'variant_norm': car.cars_standard.variant_norm,
                'model_group_norm': car.cars_standard.model_group_norm,
            }
        
        # Get category info
        category_info = None
        if car.category:
            category_info = {
                'name': car.category.name,
                'id': car.category.id
            }
        
        # Parse images if available
        images = []
        if car.images:
            try:
                import json
                images = json.loads(car.images) if isinstance(car.images, str) else car.images
            except:
                images = []
        
        data = {
            'id': car.id,
            'source': dict(CarUnified.SOURCE_CHOICES).get(car.source, car.source),
            'listing_url': car.listing_url,
            'brand': car.brand,
            'model': car.model,
            'model_group': car.model_group,
            'variant': car.variant,
            'condition': car.condition,
            'year': car.year,
            'mileage': car.mileage,
            'transmission': car.transmission,
            'seat_capacity': car.seat_capacity,
            'engine_cc': car.engine_cc,
            'fuel_type': car.fuel_type,
            'price': car.price,
            'location': car.location,
            'information_ads': car.information_ads,
            'images': images,
            'status': car.status,
            'ads_tag': car.ads_tag,
            'is_deleted': car.is_deleted,
            'last_scraped_at': car.last_scraped_at.strftime('%Y-%m-%d %H:%M:%S') if car.last_scraped_at else None,
            'version': car.version,
            'sold_at': car.sold_at.strftime('%Y-%m-%d %H:%M:%S') if car.sold_at else None,
            'last_status_check': car.last_status_check.strftime('%Y-%m-%d %H:%M:%S') if car.last_status_check else None,
            'information_ads_date': car.information_ads_date.strftime('%Y-%m-%d') if car.information_ads_date else None,
            'created_at': car.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'updated_at': car.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
            'standard_info': standard_info,
            'category_info': category_info
        }
        
        return JsonResponse({
            'success': True,
            'data': data
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# Verified Phones Management
@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def verified_phones_view(request):
    """Verified phones management view with DataTables"""
    from django.utils import timezone
    
    context = {
        'page_title': 'Verified Phones',
        'total_phones': VerifiedPhone.objects.count(),
        'active_phones': VerifiedPhone.objects.filter(is_active=True).count(),
        'expired_phones': VerifiedPhone.objects.filter(is_active=True).filter(
            verified_at__lt=timezone.now() - timezone.timedelta(days=30)
        ).count(),
        'today_verifications': VerifiedPhone.objects.filter(
            verified_at__date=timezone.now().date()
        ).count(),
    }
    return render(request, 'admin/verified-phones.html', context)

@csrf_exempt
@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def verified_phones_api(request):
    """API endpoint for DataTables verified phones data"""
    try:
        # Check if export is requested
        export_format = request.GET.get('export')
        if export_format in ['csv', 'excel']:
            return export_verified_phones(request, export_format)
        
        # DataTables parameters
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '').strip()
        
        # Ordering
        order_column_index = int(request.GET.get('order[0][column]', 0))
        order_direction = request.GET.get('order[0][dir]', 'asc')
        
        # Column mapping for ordering
        columns = ['id', 'phone_number', 'verified_at', 'last_accessed', 'access_count', 'is_active', 'ip_address']
        order_column = columns[order_column_index] if order_column_index < len(columns) else 'id'
        
        if order_direction == 'desc':
            order_column = '-' + order_column
        
        # Base queryset
        queryset = VerifiedPhone.objects.all()
        
        # Additional filtering
        status_filter = request.GET.get('status_filter')
        if status_filter == 'active':
            queryset = queryset.filter(is_active=True)
        elif status_filter == 'inactive':
            queryset = queryset.filter(is_active=False)
        elif status_filter == 'expired':
            from django.utils import timezone
            queryset = queryset.filter(
                is_active=True,
                verified_at__lt=timezone.now() - timezone.timedelta(days=30)
            )
        
        # Search filtering
        if search_value:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(phone_number__icontains=search_value) |
                Q(ip_address__icontains=search_value) |
                Q(user_agent__icontains=search_value)
            )
        
        # Total records
        total_records = VerifiedPhone.objects.count()
        filtered_records = queryset.count()
        
        # Apply ordering and pagination
        queryset = queryset.order_by(order_column)[start:start + length]
        
        # Build data for DataTables
        data = []
        for phone in queryset:
            from django.utils import timezone
            
            # Check if expired
            is_expired = phone.is_expired()
            
            # Format status
            if not phone.is_active:
                status_badge = '<span class="badge badge-error">Inactive</span>'
            elif is_expired:
                status_badge = '<span class="badge badge-warning">Expired</span>'
            else:
                status_badge = '<span class="badge badge-success">Active</span>'
            
            # Format access count with badge
            access_count_formatted = f'<span class="badge badge-outline">{phone.access_count}</span>'
            
            # Format phone number (mask middle digits for privacy)
            masked_phone = phone.phone_number[:4] + '*' * (len(phone.phone_number) - 8) + phone.phone_number[-4:]
            
            # Actions column
            actions = f'''<div class="btn-group btn-group-sm" role="group">
                <button type="button" class="btn btn-info btn-sm" onclick="viewPhoneDetails({phone.id})" title="View Details">
                    <i class="fas fa-eye"></i>
                </button>
                <button type="button" class="btn btn-warning btn-sm" onclick="togglePhoneStatus({phone.id}, {str(phone.is_active).lower()})" title="Toggle Status">
                    <i class="fas fa-toggle-{'on' if phone.is_active else 'off'}"></i>
                </button>
            </div>'''
            
            data.append([
                phone.id,
                masked_phone,
                phone.verified_at.strftime('%Y-%m-%d %H:%M'),
                phone.last_accessed.strftime('%Y-%m-%d %H:%M'),
                access_count_formatted,
                status_badge,
                phone.ip_address or '-',
                actions
            ])
        
        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': filtered_records,
            'data': data
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def export_verified_phones(request, export_format):
    """Export verified phones data in CSV or Excel format"""
    try:
        import csv
        from django.utils import timezone
        from datetime import datetime
        
        # Build queryset with filters
        queryset = VerifiedPhone.objects.all()
        
        # Apply filters
        status_filter = request.GET.get('status_filter')
        if status_filter == 'active':
            queryset = queryset.filter(is_active=True)
        elif status_filter == 'inactive':
            queryset = queryset.filter(is_active=False)
        elif status_filter == 'expired':
            queryset = queryset.filter(
                is_active=True,
                verified_at__lt=timezone.now() - timezone.timedelta(days=30)
            )
        
        search_value = request.GET.get('search[value]', '').strip()
        if search_value:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(phone_number__icontains=search_value) |
                Q(ip_address__icontains=search_value) |
                Q(user_agent__icontains=search_value)
            )
        
        queryset = queryset.order_by('-id')
        
        if export_format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="verified_phones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
            
            writer = csv.writer(response)
            writer.writerow(['ID', 'Phone Number', 'Verified At', 'Last Accessed', 'Access Count', 'Status', 'IP Address', 'User Agent'])
            
            for phone in queryset:
                is_expired = phone.is_expired()
                if not phone.is_active:
                    status = 'Inactive'
                elif is_expired:
                    status = 'Expired'
                else:
                    status = 'Active'
                
                writer.writerow([
                    phone.id,
                    phone.phone_number,
                    phone.verified_at.strftime('%Y-%m-%d %H:%M:%S'),
                    phone.last_accessed.strftime('%Y-%m-%d %H:%M:%S'),
                    phone.access_count,
                    status,
                    phone.ip_address or '',
                    phone.user_agent or ''
                ])
            
            return response
            
        elif export_format == 'excel':
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill
                from io import BytesIO
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Verified Phones"
                
                # Headers
                headers = ['ID', 'Phone Number', 'Verified At', 'Last Accessed', 'Access Count', 'Status', 'IP Address', 'User Agent']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                # Data
                for row, phone in enumerate(queryset, 2):
                    is_expired = phone.is_expired()
                    if not phone.is_active:
                        status = 'Inactive'
                    elif is_expired:
                        status = 'Expired'
                    else:
                        status = 'Active'
                    
                    ws.cell(row=row, column=1, value=phone.id)
                    ws.cell(row=row, column=2, value=phone.phone_number)
                    ws.cell(row=row, column=3, value=phone.verified_at.strftime('%Y-%m-%d %H:%M:%S'))
                    ws.cell(row=row, column=4, value=phone.last_accessed.strftime('%Y-%m-%d %H:%M:%S'))
                    ws.cell(row=row, column=5, value=phone.access_count)
                    ws.cell(row=row, column=6, value=status)
                    ws.cell(row=row, column=7, value=phone.ip_address or '')
                    ws.cell(row=row, column=8, value=phone.user_agent or '')
                
                # Auto-size columns
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
                
                # Save to BytesIO
                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                
                response = HttpResponse(
                    buffer.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="verified_phones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
                
                return response
                
            except ImportError:
                # Fallback to CSV if openpyxl is not available
                return export_verified_phones(request, 'csv')
        
    except Exception as e:
        return JsonResponse({'error': f'Export failed: {str(e)}'}, status=500)

@csrf_exempt
@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def verified_phone_detail_api(request, phone_id):
    """API endpoint to get detailed phone information"""
    try:
        phone = get_object_or_404(VerifiedPhone, id=phone_id)
        
        data = {
            'id': phone.id,
            'phone_number': phone.phone_number,
            'verified_at': phone.verified_at.strftime('%Y-%m-%d %H:%M:%S'),
            'last_accessed': phone.last_accessed.strftime('%Y-%m-%d %H:%M:%S'),
            'access_count': phone.access_count,
            'is_active': phone.is_active,
            'is_expired': phone.is_expired(),
            'user_agent': phone.user_agent,
            'ip_address': phone.ip_address,
        }
        
        return JsonResponse({
            'success': True,
            'data': data
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def toggle_phone_status(request, phone_id):
    """Toggle phone active status"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=400)
        
    try:
        phone = get_object_or_404(VerifiedPhone, id=phone_id)
        phone.is_active = not phone.is_active
        phone.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Phone status updated to {"Active" if phone.is_active else "Inactive"}',
            'is_active': phone.is_active
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# OTP Sessions Management
@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def otp_sessions_view(request):
    """OTP sessions management view with DataTables"""
    from django.utils import timezone
    
    context = {
        'page_title': 'OTP Sessions',
        'total_sessions': OTPSession.objects.count(),
        'used_sessions': OTPSession.objects.filter(is_used=True).count(),
        'active_sessions': OTPSession.objects.filter(is_used=False).count(),
        'today_sessions': OTPSession.objects.filter(
            created_at__date=timezone.now().date()
        ).count(),
    }
    return render(request, 'admin/otp-sessions.html', context)

@csrf_exempt
@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def otp_sessions_api(request):
    """API endpoint for DataTables OTP sessions data"""
    try:
        # Check if export is requested
        export_format = request.GET.get('export')
        if export_format in ['csv', 'excel']:
            return export_otp_sessions(request, export_format)
        
        # DataTables parameters
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '').strip()
        
        # Ordering
        order_column_index = int(request.GET.get('order[0][column]', 0))
        order_direction = request.GET.get('order[0][dir]', 'asc')
        
        # Column mapping for ordering
        columns = ['id', 'phone_number', 'verification_id', 'created_at', 'is_used', 'ip_address']
        order_column = columns[order_column_index] if order_column_index < len(columns) else 'id'
        
        if order_direction == 'desc':
            order_column = '-' + order_column
        
        # Base queryset
        queryset = OTPSession.objects.all()
        
        # Additional filtering
        status_filter = request.GET.get('status_filter')
        if status_filter == 'used':
            queryset = queryset.filter(is_used=True)
        elif status_filter == 'unused':
            queryset = queryset.filter(is_used=False)
        elif status_filter == 'expired':
            from django.utils import timezone
            queryset = queryset.filter(
                is_used=False,
                created_at__lt=timezone.now() - timezone.timedelta(minutes=1)
            )
        
        # Search filtering
        if search_value:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(phone_number__icontains=search_value) |
                Q(otp_code__icontains=search_value) |
                Q(ip_address__icontains=search_value)
            )
        
        # Total records
        total_records = OTPSession.objects.count()
        filtered_records = queryset.count()
        
        # Apply ordering and pagination
        queryset = queryset.order_by(order_column)[start:start + length]
        
        # Build data for DataTables
        data = []
        for otp in queryset:
            from django.utils import timezone
            
            # Check if expired
            is_expired = otp.is_expired()
            
            # Format status
            if otp.is_used:
                status_badge = '<span class="badge badge-success">Used</span>'
            elif is_expired:
                status_badge = '<span class="badge badge-error">Expired</span>'
            else:
                status_badge = '<span class="badge badge-warning">Active</span>'
            
            # Format phone number (mask middle digits for privacy)
            masked_phone = otp.phone_number[:4] + '*' * (len(otp.phone_number) - 8) + otp.phone_number[-4:]
            
            # Format verification ID (partially masked for security)
            masked_verification_id = (otp.verification_id[:3] + '***' + otp.verification_id[-3:]) if otp.verification_id and len(otp.verification_id) > 6 else (otp.verification_id or '-')
            
            # Actions column
            actions = f'''<div class="btn-group btn-group-sm" role="group">
                <button type="button" class="btn btn-info btn-sm" onclick="viewOTPDetails({otp.id})" title="View Details">
                    <i class="fas fa-eye"></i>
                </button>
            </div>'''
            
            data.append([
                otp.id,
                masked_phone,
                masked_verification_id,
                otp.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                status_badge,
                otp.ip_address or '-',
                actions
            ])
        
        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': filtered_records,
            'data': data
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def export_otp_sessions(request, export_format):
    """Export OTP sessions data in CSV or Excel format"""
    try:
        import csv
        from django.utils import timezone
        from datetime import datetime
        
        # Build queryset with filters
        queryset = OTPSession.objects.all()
        
        # Apply filters
        status_filter = request.GET.get('status_filter')
        if status_filter == 'used':
            queryset = queryset.filter(is_used=True)
        elif status_filter == 'unused':
            queryset = queryset.filter(is_used=False)
        elif status_filter == 'expired':
            queryset = queryset.filter(
                is_used=False,
                created_at__lt=timezone.now() - timezone.timedelta(minutes=1)
            )
        
        search_value = request.GET.get('search[value]', '').strip()
        if search_value:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(phone_number__icontains=search_value) |
                Q(otp_code__icontains=search_value) |
                Q(ip_address__icontains=search_value)
            )
        
        queryset = queryset.order_by('-id')
        
        if export_format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="otp_sessions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
            
            writer = csv.writer(response)
            writer.writerow(['ID', 'Phone Number', 'OTP Code', 'Verification ID', 'Created At', 'Expires At', 'Status', 'IP Address'])
            
            for otp in queryset:
                is_expired = otp.is_expired()
                
                if otp.is_used:
                    status = 'Used'
                elif is_expired:
                    status = 'Expired'
                else:
                    status = 'Active'
                
                # Mask OTP code for security
                masked_otp = otp.otp_code[:2] + '*' * (len(otp.otp_code) - 4) + otp.otp_code[-2:] if otp.otp_code else ''
                
                # Mask phone number
                masked_phone = otp.phone_number[:4] + '*' * (len(otp.phone_number) - 8) + otp.phone_number[-4:] if otp.phone_number else ''
                
                writer.writerow([
                    otp.id,
                    masked_phone,
                    masked_otp,
                    otp.verification_id or '',
                    otp.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    otp.expires_at.strftime('%Y-%m-%d %H:%M:%S') if otp.expires_at else '',
                    status,
                    otp.ip_address or ''
                ])
            
            return response
            
        elif export_format == 'excel':
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill
                from io import BytesIO
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "OTP Sessions"
                
                # Headers
                headers = ['ID', 'Phone Number', 'OTP Code', 'Verification ID', 'Created At', 'Expires At', 'Status', 'IP Address']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                # Data
                for row, otp in enumerate(queryset, 2):
                    is_expired = otp.is_expired()
                    
                    if otp.is_used:
                        status = 'Used'
                    elif is_expired:
                        status = 'Expired'
                    else:
                        status = 'Active'
                    
                    # Mask OTP code for security
                    masked_otp = otp.otp_code[:2] + '*' * (len(otp.otp_code) - 4) + otp.otp_code[-2:] if otp.otp_code else ''
                    
                    # Mask phone number
                    masked_phone = otp.phone_number[:4] + '*' * (len(otp.phone_number) - 8) + otp.phone_number[-4:] if otp.phone_number else ''
                    
                    ws.cell(row=row, column=1, value=otp.id)
                    ws.cell(row=row, column=2, value=masked_phone)
                    ws.cell(row=row, column=3, value=masked_otp)
                    ws.cell(row=row, column=4, value=otp.verification_id or '')
                    ws.cell(row=row, column=5, value=otp.created_at.strftime('%Y-%m-%d %H:%M:%S'))
                    ws.cell(row=row, column=6, value=otp.expires_at.strftime('%Y-%m-%d %H:%M:%S') if otp.expires_at else '')
                    ws.cell(row=row, column=7, value=status)
                    ws.cell(row=row, column=8, value=otp.ip_address or '')
                
                # Auto-size columns
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
                
                # Save to BytesIO
                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                
                response = HttpResponse(
                    buffer.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="otp_sessions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
                
                return response
                
            except ImportError:
                # Fallback to CSV if openpyxl is not available
                return export_otp_sessions(request, 'csv')
        
    except Exception as e:
        return JsonResponse({'error': f'Export failed: {str(e)}'}, status=500)

@csrf_exempt
@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def otp_session_detail_api(request, session_id):
    """API endpoint to get detailed OTP session information"""
    try:
        otp = get_object_or_404(OTPSession, id=session_id)
        
        data = {
            'id': otp.id,
            'phone_number': otp.phone_number,
            'verification_id': otp.verification_id,
            'created_at': otp.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'is_used': otp.is_used,
            'is_expired': otp.is_expired(),
            'is_valid': otp.is_valid(),
            'ip_address': otp.ip_address,
            'expires_at': (otp.created_at + timezone.timedelta(minutes=1)).strftime('%Y-%m-%d %H:%M:%S'),
        }
        
        return JsonResponse({
            'success': True,
            'data': data
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# Categories Management (Brand Categories System)
@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def categories_management_view(request):
    """Categories management view with CRUD operations"""
    from django.db.models import Count
    
    # Get categories with brand counts
    categories = Category.objects.annotate(
        brand_count=Count('brandcategory')
    ).order_by('name')
    
    # Get unclassified brands count
    classified_brands = BrandCategory.objects.values_list('brand', flat=True)
    all_brands = CarUnified.objects.values_list('brand', flat=True).distinct()
    unclassified_count = len(set(all_brands) - set(classified_brands))
    
    context = {
        'page_title': 'Brand Categories Management',
        'categories': categories,
        'total_categories': categories.count(),
        'total_classified_brands': BrandCategory.objects.count(),
        'unclassified_brands': unclassified_count,
        'total_unique_brands': CarUnified.objects.values('brand').distinct().count(),
    }
    return render(request, 'admin/categories-management.html', context)

@csrf_exempt
@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def category_create(request):
    """Create new category"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=400)
    
    try:
        data = json.loads(request.body)
        name = data.get('name', '').strip()
        reduction_percentage = data.get('reduction_percentage', 0.0)
        
        if not name:
            return JsonResponse({'error': 'Category name is required'}, status=400)
        
        # Validate reduction percentage
        try:
            reduction_percentage = float(reduction_percentage)
            if reduction_percentage < 0 or reduction_percentage > 100:
                return JsonResponse({'error': 'Reduction percentage must be between 0 and 100'}, status=400)
        except (TypeError, ValueError):
            return JsonResponse({'error': 'Invalid reduction percentage'}, status=400)
        
        # Check for duplicate
        if Category.objects.filter(name=name).exists():
            return JsonResponse({'error': 'Category with this name already exists'}, status=400)
        
        category = Category.objects.create(name=name, reduction_percentage=reduction_percentage)
        
        return JsonResponse({
            'success': True,
            'message': f'Category "{category.name}" created successfully',
            'category_id': category.id,
            'category_name': category.name
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def category_edit(request, category_id):
    """Edit existing category"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=400)
    
    try:
        category = get_object_or_404(Category, id=category_id)
        data = json.loads(request.body)
        new_name = data.get('name', '').strip()
        reduction_percentage = data.get('reduction_percentage', category.reduction_percentage)
        
        if not new_name:
            return JsonResponse({'error': 'Category name is required'}, status=400)
        
        # Validate reduction percentage
        try:
            reduction_percentage = float(reduction_percentage)
            if reduction_percentage < 0 or reduction_percentage > 100:
                return JsonResponse({'error': 'Reduction percentage must be between 0 and 100'}, status=400)
        except (TypeError, ValueError):
            return JsonResponse({'error': 'Invalid reduction percentage'}, status=400)
        
        # Check for duplicate (excluding current category)
        if Category.objects.filter(name=new_name).exclude(id=category_id).exists():
            return JsonResponse({'error': 'Category with this name already exists'}, status=400)
        
        old_name = category.name
        old_reduction = category.reduction_percentage
        category.name = new_name
        category.reduction_percentage = reduction_percentage
        category.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Category updated from "{old_name}" to "{new_name}"',
            'category_id': category.id,
            'category_name': category.name
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def category_delete(request, category_id):
    """Delete category (with safety checks)"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=400)
    
    try:
        category = get_object_or_404(Category, id=category_id)
        
        # Check if category has brands assigned
        brand_count = category.brandcategory_set.count()
        if brand_count > 0:
            return JsonResponse({
                'error': f'Cannot delete category. It has {brand_count} brands assigned. Please reassign or remove the brands first.'
            }, status=400)
        
        category_name = category.name
        category.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Category "{category_name}" deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def category_brands_api(request, category_id):
    """Get brands assigned to a specific category"""
    try:
        category = get_object_or_404(Category, id=category_id)
        brand_categories = BrandCategory.objects.filter(category=category).order_by('brand')
        
        brands = []
        for bc in brand_categories:
            # Count cars for this brand
            car_count = CarUnified.objects.filter(brand=bc.brand).count()
            brands.append({
                'id': bc.id,
                'brand': bc.brand,
                'car_count': car_count,
                'created_at': bc.created_at.strftime('%Y-%m-%d %H:%M')
            })
        
        return JsonResponse({
            'success': True,
            'category': {
                'id': category.id,
                'name': category.name
            },
            'brands': brands,
            'total_brands': len(brands)
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# Brand Classification Interface
@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def brand_classification_view(request):
    """Brand classification interface"""
    from django.db.models import Count
    
    # Get all brands that exist in CarUnified
    existing_brands = set(CarUnified.objects.values_list('brand', flat=True).distinct())
    
    # Get only valid classified brands (that exist in CarUnified)
    valid_classified_brands = BrandCategory.objects.filter(brand__in=existing_brands)
    
    # Get statistics
    total_brands = len(existing_brands)
    classified_brands = valid_classified_brands.count()
    unclassified_brands = total_brands - classified_brands
    total_categories = Category.objects.count()
    
    context = {
        'page_title': 'Brand Classification',
        'total_brands': total_brands,
        'classified_brands': classified_brands,
        'unclassified_brands': unclassified_brands,
        'total_categories': total_categories,
        'categories': Category.objects.all().order_by('name'),
    }
    return render(request, 'admin/brand-classification.html', context)

@csrf_exempt
@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def brands_data_api(request):
    """API for brand classification DataTables"""
    try:
        # DataTables parameters
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '').strip()
        
        # Filter parameters
        status_filter = request.GET.get('status_filter', '')  # classified, unclassified, all
        category_filter = request.GET.get('category_filter', '')
        
        # Get all unique brands with car counts
        from django.db.models import Count, Q
        
        # Base query: get all brands with their car counts
        brands_query = CarUnified.objects.values('brand').annotate(
            car_count=Count('id')
        )
        
        # Get all brands that exist in CarUnified
        existing_brands = set(CarUnified.objects.values_list('brand', flat=True).distinct())
        
        # Get classified brands mapping (only for brands that exist in CarUnified)
        brand_categories_map = {}
        for bc in BrandCategory.objects.select_related('category').filter(brand__in=existing_brands):
            brand_categories_map[bc.brand] = {
                'category_id': bc.category.id,
                'category_name': bc.category.name,
                'mapping_id': bc.id
            }
        
        # Apply search filter
        if search_value:
            brands_query = brands_query.filter(brand__icontains=search_value)
        
        # Apply status filter
        if status_filter == 'classified':
            classified_brand_names = list(brand_categories_map.keys())
            brands_query = brands_query.filter(brand__in=classified_brand_names)
        elif status_filter == 'unclassified':
            classified_brand_names = list(brand_categories_map.keys())
            brands_query = brands_query.exclude(brand__in=classified_brand_names)
        
        # Apply category filter
        if category_filter:
            category_brands = BrandCategory.objects.filter(
                category_id=category_filter
            ).values_list('brand', flat=True)
            brands_query = brands_query.filter(brand__in=category_brands)
        
        # Get total counts
        total_records = CarUnified.objects.values('brand').distinct().count()
        filtered_records = brands_query.count()
        
        # Apply pagination and ordering
        brands_query = brands_query.order_by('brand')[start:start + length]
        
        # Build data for DataTables
        data = []
        for brand_data in brands_query:
            brand_name = brand_data['brand']
            car_count = brand_data['car_count']
            
            # Check if brand is classified
            category_info = brand_categories_map.get(brand_name)
            
            if category_info:
                # Classified
                status_html = f'<span class="badge badge-success">Classified</span>'
                category_html = f'<span class="badge badge-primary">{category_info["category_name"]}</span>'
                action_btn = f'''<button class="btn btn-sm btn-warning" onclick="reassignBrand('{brand_name}', {category_info['category_id']}, {category_info['mapping_id']})" title="Reassign">
                    <i class="fas fa-edit"></i>
                </button>
                <button class="btn btn-sm btn-error" onclick="removeBrandClassification('{brand_name}', {category_info['mapping_id']})" title="Remove">
                    <i class="fas fa-trash"></i>
                </button>'''
            else:
                # Unclassified
                status_html = f'<span class="badge badge-error">Unclassified</span>'
                category_html = '<span class="text-base-content/40">-</span>'
                action_btn = f'''<button class="btn btn-sm btn-primary" onclick="assignBrand('{brand_name}')" title="Assign">
                    <i class="fas fa-plus"></i> Assign
                </button>'''
            
            data.append([
                brand_name,
                f'{car_count:,}',
                status_html,
                category_html,
                action_btn
            ])
        
        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': filtered_records,
            'data': data
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def assign_brand_to_category(request):
    """Assign a brand to a category"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=400)
    
    try:
        data = json.loads(request.body)
        brand_name = data.get('brand_name', '').strip()
        category_id = data.get('category_id')
        
        if not brand_name or not category_id:
            return JsonResponse({'error': 'Brand name and category are required'}, status=400)
        
        # Validate category exists
        category = get_object_or_404(Category, id=category_id)
        
        # Validate brand exists in cars_unified
        if not CarUnified.objects.filter(brand=brand_name).exists():
            return JsonResponse({'error': 'Brand not found in database'}, status=400)
        
        # Check if brand is already classified
        if BrandCategory.objects.filter(brand=brand_name).exists():
            return JsonResponse({'error': 'Brand is already classified'}, status=400)
        
        # Create the brand-category mapping
        brand_category = BrandCategory.objects.create(
            brand=brand_name,
            category=category
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Brand "{brand_name}" assigned to category "{category.name}"',
            'brand_category_id': brand_category.id
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
@csrf_exempt
@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def reassign_brand_to_category(request):
    """Reassign a brand to a different category"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=400)
    
    try:
        data = json.loads(request.body)
        brand_name = data.get('brand_name', '').strip()
        new_category_id = data.get('category_id')
        mapping_id = data.get('mapping_id')
        
        if not brand_name or not new_category_id:
            return JsonResponse({'error': 'Brand name and category are required'}, status=400)
        
        # Validate new category exists
        new_category = get_object_or_404(Category, id=new_category_id)
        
        # Get existing mapping - find by brand name if mapping_id not provided
        if mapping_id:
            brand_category = get_object_or_404(BrandCategory, id=mapping_id, brand=brand_name)
        else:
            brand_category = get_object_or_404(BrandCategory, brand=brand_name)
        old_category_name = brand_category.category.name
        
        # Update the mapping
        brand_category.category = new_category
        brand_category.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Brand "{brand_name}" reassigned from "{old_category_name}" to "{new_category.name}"'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def remove_brand_classification(request):
    """Remove brand classification (make it unclassified)"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=400)
    
    try:
        data = json.loads(request.body)
        brand_name = data.get('brand', '').strip()  # Changed from brand_name to brand
        mapping_id = data.get('mapping_id')
        
        if not brand_name:
            return JsonResponse({'error': 'Brand name is required'}, status=400)
        
        # Get and delete the mapping - find by brand name only if mapping_id not provided
        if mapping_id:
            brand_category = get_object_or_404(BrandCategory, id=mapping_id, brand=brand_name)
        else:
            brand_category = get_object_or_404(BrandCategory, brand=brand_name)
        category_name = brand_category.category.name
        brand_category.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Brand "{brand_name}" removed from category "{category_name}"'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@login_required
@user_passes_test(is_staff_user, login_url='/login/')
def get_unclassified_brands_api(request):
    """Get list of unclassified brands"""
    try:
        from django.db.models import Count
        
        # Get all brands from cars_unified
        all_brands_query = CarUnified.objects.values('brand').annotate(
            car_count=Count('id')
        ).order_by('brand')
        
        # Get classified brands
        classified_brands = set(BrandCategory.objects.values_list('brand', flat=True))
        
        # Filter unclassified brands
        unclassified_brands = []
        for brand_data in all_brands_query:
            if brand_data['brand'] not in classified_brands:
                unclassified_brands.append({
                    'brand': brand_data['brand'],
                    'car_count': brand_data['car_count']
                })
        
        return JsonResponse({
            'success': True,
            'unclassified_brands': unclassified_brands,
            'total_unclassified': len(unclassified_brands)
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# Price Tiers Management Views
@login_required
def price_tiers_management_view(request):
    """Price Tiers Management View"""
    try:
        # Get all price tiers
        price_tiers = PriceTier.objects.all().order_by('order', 'min_price')
        
        # Calculate statistics
        total_tiers = price_tiers.count()
        active_tiers = price_tiers.filter(is_active=True).count()
        
        # Check for price gaps or overlaps
        has_issues = False
        issues = []
        
        active_price_tiers = price_tiers.filter(is_active=True).order_by('min_price')
        for i, tier in enumerate(active_price_tiers):
            if i > 0:
                prev_tier = active_price_tiers[i-1]
                if prev_tier.max_price and float(prev_tier.max_price) < float(tier.min_price):
                    # Gap detected (only if there's actually a gap, not just different by 1)
                    has_issues = True
                    issues.append(f"Price gap between {prev_tier.name} and {tier.name}")
                elif prev_tier.max_price and float(prev_tier.max_price) > float(tier.min_price):
                    # Overlap detected (only if max price is greater than min price of next tier)
                    has_issues = True
                    issues.append(f"Price overlap between {prev_tier.name} and {tier.name}")
        
        context = {
            'price_tiers': price_tiers,
            'total_tiers': total_tiers,
            'active_tiers': active_tiers,
            'inactive_tiers': total_tiers - active_tiers,
            'has_issues': has_issues,
            'issues': issues,
        }
        
        return render(request, 'admin/price-tiers-management.html', context)
        
    except Exception as e:
        messages.error(request, f'Error loading price tiers: {str(e)}')
        return redirect('main:categories_management_view')


@login_required
def price_tier_create(request):
    """Create new price tier"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        name = data.get('name', '').strip()
        min_price = data.get('min_price')
        max_price = data.get('max_price')
        reduction_percentage = data.get('reduction_percentage')
        
        # Validation
        if not name:
            return JsonResponse({'error': 'Tier name is required'}, status=400)
            
        if min_price is None or min_price < 0:
            return JsonResponse({'error': 'Valid minimum price is required'}, status=400)
            
        if max_price is not None and max_price <= min_price:
            return JsonResponse({'error': 'Maximum price must be greater than minimum price'}, status=400)
            
        if reduction_percentage is None or reduction_percentage < 0 or reduction_percentage > 100:
            return JsonResponse({'error': 'Reduction percentage must be between 0-100%'}, status=400)
        
        # Check for existing tier with same name
        if PriceTier.objects.filter(name=name).exists():
            return JsonResponse({'error': 'Price tier with this name already exists'}, status=400)
        
        # Create new tier
        tier = PriceTier.objects.create(
            name=name,
            min_price=min_price,
            max_price=max_price,
            reduction_percentage=reduction_percentage,
            order=PriceTier.objects.count() + 1
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Price tier "{name}" created successfully'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def price_tier_edit(request, tier_id):
    """Edit existing price tier"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        tier = get_object_or_404(PriceTier, id=tier_id)
        data = json.loads(request.body)
        
        name = data.get('name', '').strip()
        min_price = data.get('min_price')
        max_price = data.get('max_price')
        reduction_percentage = data.get('reduction_percentage')
        
        # Validation
        if not name:
            return JsonResponse({'error': 'Tier name is required'}, status=400)
            
        if min_price is None or min_price < 0:
            return JsonResponse({'error': 'Valid minimum price is required'}, status=400)
            
        if max_price is not None and max_price <= min_price:
            return JsonResponse({'error': 'Maximum price must be greater than minimum price'}, status=400)
            
        if reduction_percentage is None or reduction_percentage < 0 or reduction_percentage > 100:
            return JsonResponse({'error': 'Reduction percentage must be between 0-100%'}, status=400)
        
        # Check for existing tier with same name (excluding current tier)
        if PriceTier.objects.filter(name=name).exclude(id=tier_id).exists():
            return JsonResponse({'error': 'Price tier with this name already exists'}, status=400)
        
        # Update tier
        tier.name = name
        tier.min_price = min_price
        tier.max_price = max_price
        tier.reduction_percentage = reduction_percentage
        tier.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Price tier "{name}" updated successfully'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def price_tier_delete(request, tier_id):
    """Delete price tier"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        tier = get_object_or_404(PriceTier, id=tier_id)
        tier_name = tier.name
        tier.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Price tier "{tier_name}" deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)